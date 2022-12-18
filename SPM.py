# to do:
# Discharge and charge capacities don't match. Verify calculation in excel.


import numpy
import sys
import re
import time
import openpyxl as xl
from scipy.integrate import cumtrapz
import scipy.optimize
import pathlib
import scipy.sparse.linalg


def isNum(x):
    """
    Check to see if argument is a number.
    Returns true is x is a number, false if not
    """

    try:
        float(x)
        return True
    except:
        return False


def conv_numbers(s):
    """
    Convert string to number where appropriate. Also can handle arrays.
    """

    try:
        length = len(s)
    except:
        length = 0

    if length <= 1:
        try:
            q = float(s[0])
            return q
        except ValueError:
            return s[0]
    else:
        try:
            q = [float(val) for val in s]
            return q
        except ValueError:
            return s


def density(material):
    """
    Density of pure phase material [kg/m^3]
    """

    result = {
        "NMC": lambda: 4750.0,
        "LCO": lambda: 5031.0,
        "NCA": lambda: 4450.0,
        "LMnO": lambda: 4290.0,
        "MCMB1": lambda: 2260.0,
        "MCMB2": lambda: 2260.0,
        "PVDF": lambda: 1760.0,
        "carbon": lambda: 2000.0,
        "separator": lambda: 1324.0,
        "electrolyte": lambda: 1204.0,
        "Alfoil": lambda: 2700.0,
        "Cufoil": lambda: 8930.0,
    }[material]()
    return result


def thermodynamicCapacity(material):
    """
    Thermodynamic capacity of material (Ah/g)?
    """

    result = {
        "NMC": lambda: 277.54,
        "LCO": lambda: 273.8,
        "NCA": lambda: 278.01,
        "LMnO": lambda: 148.22,
        "MCMB1": lambda: 363.0,
        "MCMB2": lambda: 363.0,
    }[material]()
    return result


def returnDs(y, material):
    """
    Returns the lithium diffusion coefficient as a function of y in Liy[M]O4
    """

    if material == "NMC":
        """data from Shaju2004a"""
        return (
            (
                numpy.power(
                    10,
                    -63.061 * numpy.power(y, 5)
                    + 239.8 * numpy.power(y, 4)
                    - 343.74 * numpy.power(y, 3)
                    + 232.45 * numpy.power(y, 2)
                    - 74.337 * y
                    - 0.2517,
                )
            )
            / 100
            / 100
        )


# -----------------------------------------------


def dUdT(material, soc):
    """
    Return dU(soc)/dT for various materials
    I don't think nmc is correct (Jeon 2011)
    lco (Jeon2011)
    """
    lmno = numpy.array(
        [
            [
                5.00000000e-03,
                7.00000000e-03,
                1.00000000e-02,
                1.20000000e-02,
                1.60000000e-02,
                1.70000000e-02,
                2.40000000e-02,
                2.80000000e-02,
                3.10000000e-02,
                3.30000000e-02,
                3.70000000e-02,
                3.80000000e-02,
                4.00000000e-02,
                4.20000000e-02,
                4.30000000e-02,
                4.50000000e-02,
                4.70000000e-02,
                4.90000000e-02,
                5.00000000e-02,
                5.40000000e-02,
                5.60000000e-02,
                5.90000000e-02,
                6.10000000e-02,
                6.40000000e-02,
                7.00000000e-02,
                7.70000000e-02,
                8.70000000e-02,
                9.90000000e-02,
                1.10000000e-01,
                1.22000000e-01,
                1.34000000e-01,
                1.46000000e-01,
                1.65000000e-01,
                1.81000000e-01,
                1.93000000e-01,
                2.05000000e-01,
                2.17000000e-01,
                2.31000000e-01,
                2.42000000e-01,
                2.57000000e-01,
                2.92000000e-01,
                3.20000000e-01,
                3.51000000e-01,
                3.79000000e-01,
                4.03000000e-01,
                4.16000000e-01,
                4.28000000e-01,
                4.38000000e-01,
                4.49000000e-01,
                4.68000000e-01,
                4.78000000e-01,
                4.87000000e-01,
                4.97000000e-01,
                5.04000000e-01,
                5.17000000e-01,
                5.32000000e-01,
                7.98000000e-01,
            ],
            [
                2.50000000e-05,
                3.40000000e-05,
                4.60000000e-05,
                6.10000000e-05,
                7.80000000e-05,
                9.90000000e-05,
                1.36000000e-04,
                1.68000000e-04,
                1.96000000e-04,
                2.10000000e-04,
                2.28000000e-04,
                2.40000000e-04,
                2.45000000e-04,
                2.51000000e-04,
                2.54000000e-04,
                2.56000000e-04,
                2.54100000e-04,
                2.49000000e-04,
                2.44000000e-04,
                2.33000000e-04,
                2.24000000e-04,
                2.07000000e-04,
                1.82000000e-04,
                1.57000000e-04,
                1.20000000e-04,
                8.70000000e-05,
                4.30000000e-05,
                8.00000000e-06,
                -1.30000000e-05,
                -3.30000000e-05,
                -4.70000000e-05,
                -5.90000000e-05,
                -7.50000000e-05,
                -9.40000000e-05,
                -1.08000000e-04,
                -1.28000000e-04,
                -1.42000000e-04,
                -1.58000000e-04,
                -1.67000000e-04,
                -1.74000000e-04,
                -1.70000000e-04,
                -1.68000000e-04,
                -1.72000000e-04,
                -1.74000000e-04,
                -1.72200000e-04,
                -1.67000000e-04,
                -1.60000000e-04,
                -1.52000000e-04,
                -1.45000000e-04,
                -1.37000000e-04,
                -1.33000000e-04,
                -1.26000000e-04,
                -1.15000000e-04,
                -1.07000000e-04,
                -1.01000000e-04,
                -1.00000000e-04,
                -1.10000000e-04,
            ],
        ]
    )
    nmc = numpy.array(
        [
            [
                0.00000000e00,
                2.00000000e-02,
                5.50000000e-02,
                9.00000000e-02,
                1.22000000e-01,
                1.66000000e-01,
                2.18000000e-01,
                2.47000000e-01,
                2.99000000e-01,
                3.49000000e-01,
                3.78000000e-01,
                4.19000000e-01,
                5.00000000e-01,
                5.76000000e-01,
                6.28000000e-01,
                7.01000000e-01,
                7.50000000e-01,
                8.05000000e-01,
                8.75000000e-01,
                9.36000000e-01,
                1.00000000e00,
            ],
            [
                -1.10598000e-04,
                -1.02285000e-04,
                -9.39628000e-05,
                -9.05218000e-05,
                -8.96616000e-05,
                -8.93714000e-05,
                -8.96616000e-05,
                -8.87910000e-05,
                -8.59305000e-05,
                -8.04788000e-05,
                -7.73177000e-05,
                -7.15863000e-05,
                -6.06830000e-05,
                -5.40809000e-05,
                -5.20703000e-05,
                -5.35109000e-05,
                -5.66617000e-05,
                -6.09628000e-05,
                -6.69949000e-05,
                -7.30165000e-05,
                -7.96186000e-05,
            ],
        ]
    )
    lco = numpy.array(
        [
            [
                5.01000000e-01,
                5.06000000e-01,
                5.12000000e-01,
                5.19000000e-01,
                5.26000000e-01,
                5.31000000e-01,
                5.34000000e-01,
                5.40000000e-01,
                5.51000000e-01,
                5.53000000e-01,
                5.57000000e-01,
                5.63000000e-01,
                5.69000000e-01,
                5.82000000e-01,
                5.85000000e-01,
                6.12000000e-01,
                6.46000000e-01,
                6.65000000e-01,
                7.07000000e-01,
                7.19000000e-01,
                7.71000000e-01,
                8.10000000e-01,
                8.47000000e-01,
                8.81000000e-01,
                9.03000000e-01,
                9.28000000e-01,
                9.57000000e-01,
                9.68000000e-01,
                9.81000000e-01,
                9.94000000e-01,
                9.97000000e-01,
                9.88000000e-01,
                9.94000000e-01,
                1.00000000e00,
            ],
            [
                -3.12214000e-04,
                -3.94787000e-04,
                -3.94787000e-04,
                -3.29419000e-04,
                -1.98684000e-04,
                -8.85941000e-05,
                2.15059000e-05,
                1.24714000e-04,
                1.59123000e-04,
                1.31596000e-04,
                6.96689000e-05,
                8.60237000e-07,
                -9.89169000e-05,
                -1.50521000e-04,
                -1.74597000e-04,
                -2.12448000e-04,
                -2.57170000e-04,
                -2.77815000e-04,
                -3.32860000e-04,
                -3.50065000e-04,
                -4.56713000e-04,
                -5.39286000e-04,
                -5.70254000e-04,
                -5.97772000e-04,
                -6.11535000e-04,
                -6.14976000e-04,
                -6.14976000e-04,
                -5.25522000e-04,
                -4.53273000e-04,
                -4.56713000e-04,
                -4.70477000e-04,
                -4.56713000e-04,
                -4.63595000e-04,
                -4.70477000e-04,
            ],
        ]
    )
    graphite = numpy.array(
        [
            [
                5.00000000e-03,
                7.00000000e-03,
                1.00000000e-02,
                1.20000000e-02,
                1.60000000e-02,
                1.70000000e-02,
                2.40000000e-02,
                2.80000000e-02,
                3.10000000e-02,
                3.30000000e-02,
                3.70000000e-02,
                3.80000000e-02,
                4.00000000e-02,
                4.20000000e-02,
                4.30000000e-02,
                4.50000000e-02,
                4.70000000e-02,
                4.90000000e-02,
                5.00000000e-02,
                5.40000000e-02,
                5.60000000e-02,
                5.90000000e-02,
                6.10000000e-02,
                6.40000000e-02,
                7.00000000e-02,
                7.70000000e-02,
                8.70000000e-02,
                9.90000000e-02,
                1.10000000e-01,
                1.22000000e-01,
                1.34000000e-01,
                1.46000000e-01,
                1.65000000e-01,
                1.81000000e-01,
                1.93000000e-01,
                2.05000000e-01,
                2.17000000e-01,
                2.31000000e-01,
                2.42000000e-01,
                2.57000000e-01,
                2.92000000e-01,
                3.20000000e-01,
                3.51000000e-01,
                3.79000000e-01,
                4.03000000e-01,
                4.16000000e-01,
                4.28000000e-01,
                4.38000000e-01,
                4.49000000e-01,
                4.68000000e-01,
                4.78000000e-01,
                4.87000000e-01,
                4.97000000e-01,
                5.04000000e-01,
                5.17000000e-01,
                5.32000000e-01,
                7.98000000e-01,
            ],
            [
                2.50000000e-05,
                3.40000000e-05,
                4.60000000e-05,
                6.10000000e-05,
                7.80000000e-05,
                9.90000000e-05,
                1.36000000e-04,
                1.68000000e-04,
                1.96000000e-04,
                2.10000000e-04,
                2.28000000e-04,
                2.40000000e-04,
                2.45000000e-04,
                2.51000000e-04,
                2.54000000e-04,
                2.56000000e-04,
                2.54100000e-04,
                2.49000000e-04,
                2.44000000e-04,
                2.33000000e-04,
                2.24000000e-04,
                2.07000000e-04,
                1.82000000e-04,
                1.57000000e-04,
                1.20000000e-04,
                8.70000000e-05,
                4.30000000e-05,
                8.00000000e-06,
                -1.30000000e-05,
                -3.30000000e-05,
                -4.70000000e-05,
                -5.90000000e-05,
                -7.50000000e-05,
                -9.40000000e-05,
                -1.08000000e-04,
                -1.28000000e-04,
                -1.42000000e-04,
                -1.58000000e-04,
                -1.67000000e-04,
                -1.74000000e-04,
                -1.70000000e-04,
                -1.68000000e-04,
                -1.72000000e-04,
                -1.74000000e-04,
                -1.72200000e-04,
                -1.67000000e-04,
                -1.60000000e-04,
                -1.52000000e-04,
                -1.45000000e-04,
                -1.37000000e-04,
                -1.33000000e-04,
                -1.26000000e-04,
                -1.15000000e-04,
                -1.07000000e-04,
                -1.01000000e-04,
                -1.00000000e-04,
                -1.10000000e-04,
            ],
        ]
    )

    result = {
        "LMnO": lambda: numpy.interp(soc, lmno[0, :], lmno[1, :]),
        "NMC": lambda: numpy.interp(soc, nmc[0, :], nmc[1, :]),
        "LCO": lambda: numpy.interp(soc, lco[0, :], lco[1, :]),
        "MCMB1": lambda: numpy.interp(soc, graphite[0, :], graphite[1, :]),
        "MCMB2": lambda: numpy.interp(soc, graphite[0, :], graphite[1, :]),
    }[material]()

    return result


def Eref(material, soc, T, fudge=0):
    """returns the Eref as function of soc of electrode
    fudge is a small fudge factor: NMC -0.01, MCMB2 0.0229
    """

    tempEffect = dUdT(material, soc)

    def NMC(soc):
        x = numpy.array(
            [
                3.125766183885334,
                -5.763342952859494,
                2.124454409988303,
                0.51324615231389,
                -1.999566459156232,
                0.992457700580878,
                0.007158221269832,
            ]
        )
        return (
            x[0] * numpy.power(soc, 3) + x[1] * numpy.power(soc, 2) + x[2] * soc + x[3]
        ) / (numpy.power(soc, 3) + x[4] * numpy.power(soc, 2) + x[5] * soc + x[6])

    def NCA(soc):
        x = numpy.array(
            [
                -47.09304396,
                -2.64754588,
                -71.38533023,
                -800.63158077,
                -340.36135014,
                1284.86452671,
                -429.83691774,
                204.34231731,
                233.74192887,
            ]
        )
        return (
            x[0] * numpy.power(soc, 5)
            + x[1] * numpy.power(soc, 4)
            + x[2] * numpy.power(soc, 3)
            + x[3] * numpy.power(soc, 2)
            + x[4] * soc
            + x[5]
        ) / (numpy.power(soc, 3) + x[6] * numpy.power(soc, 2) + x[7] * soc + x[8])

    def LCO(soc):
        if soc < 0.43:
            print("soc is too low for LCO cathode. Fixing soc=0.43\n")
            soc = 0.43
        x = numpy.array(
            [
                -4.656,
                88.669,
                -401.119,
                342.909,
                -462.471,
                433.434,
                -1,
                18.933,
                -79.532,
                37.311,
                -73.083,
                95.96,
            ]
        )
        return (
            x[0]
            + x[1] * numpy.power(soc, 2)
            + x[2] * numpy.power(soc, 4)
            + x[3] * numpy.power(soc, 6)
            + x[4] * numpy.power(soc, 8)
            + x[5] * numpy.power(soc, 10)
        ) / (
            x[6]
            + x[7] * numpy.power(soc, 2)
            + x[8] * numpy.power(soc, 4)
            + x[9] * numpy.power(soc, 6)
            + x[10] * numpy.power(soc, 8)
            + x[11] * numpy.power(soc, 10)
        )

    def MCMB1(soc):
        if Eref > 1.2:
            return 1.2
        elif Eref < 0:
            return 0.000001
        return -0.16 + 1.32 * numpy.exp(-3.0 * soc) + 10.0 * numpy.exp(-2000.0 * soc)

    def MCMB2(soc):
        if soc < 0.006:
            soc = 0.006
        x = numpy.array([0.7222, 0.1387, 0.029, -0.0172, 0.0019, 0.2808, -0.7984])
        return (
            x[0]
            + x[1] * soc
            + x[2] * numpy.sqrt(soc)
            + x[3] / soc
            + x[4] / numpy.power(soc, 1.5)
            + x[5] * numpy.exp(0.9 - 15.0 * soc)
            + x[6] * numpy.exp(0.4465 * soc - 0.4108)
        )

    Eref = {
        "NCA": lambda: NCA(soc),
        "NMC": lambda: NMC(soc),
        "LCO": lambda: LCO(soc),
        "MCMB1": lambda: MCMB1(soc),
        "MCMB2": lambda: MCMB2(soc),
    }[material]()

    return Eref + fudge + (T - 298) * tempEffect


# -----------------------------------------------


def ionicConductivity(ce, T, brugg, porosity):
    """
    Returns the effective ionic conductivity of the electrolyte (PC:EC:DEC) [S/m]
    """

    sigma = (
        0.0001
        * ce
        * numpy.power(
            -10.5
            + 0.000668 * ce
            + 0.000000494 * ce * ce
            + 0.074 * T
            - 0.0000178 * T * ce
            - 0.000000000886 * ce * ce * T
            - 0.0000696 * T * T
            + 0.000000028 * T * T * ce,
            2,
        )
    )
    return sigma * numpy.power(porosity, brugg)


# -----------------------------------------------


def storeData(data, newData):
    """Append data. If array doesn't already exist, make it."""

    try:
        data = numpy.append(data, [newData], axis=0)
    except NameError:
        data = numpy.zeros([1, numpy.shape(newData)[0]])
        data = numpy.append(data, [newData], axis=0)
    except ValueError:
        data = numpy.zeros([1, numpy.shape(newData)[0]])
        data = numpy.append(data, [newData], axis=0)

    return data


def saveData(fileName, data, headers):
    """Save data to file"""

    if sum(data[0, :]) == 0:
        data = numpy.delete(data, 0, 0)
    numpy.savetxt(fileName, data, delimiter=",", header=headers)


# -----------------------------------------------


def timeControl(
    dt,
    dt_fixed,
    voltage,
    dV,
    Iapp,
    step,
    stepIterations,
    IcutOff,
    trest,
    stepTime,
    maxV,
    minV,
    dtpulse,
    pulseTime,
    DOD,
    capacity,
    pulseSchedule,
):
    """Control the time step. Either faster or slower based on situation."""

    realdt = dt

    if step != 3:
        """Find out if the voltage is closing in on the end voltage, if so, set a warning flag"""
        if Iapp < 0 and voltage < (minV + 0.005):
            small_dt_flag = 1
            # dt = 0.01
            dt = 0.05
        # dt = 0.02
        elif (Iapp < 0 and voltage < (minV + 0.1)) or (
            Iapp > 0 and voltage > (maxV - 0.002)
        ):
            small_dt_flag = 1
            # dt = 0.1
            dt = 2.0
        elif (Iapp < 0 and voltage < (minV + 0.25)) or (
            Iapp > 0 and voltage > (maxV - 0.1)
        ):
            small_dt_flag = 1
        else:
            small_dt_flag = 0

        if stepIterations < 3 or (pulseTime <= 10 and dtpulse == 1):
            """for first couple of time steps in a cycle, keep dt small for good resolution in transition period"""
            dt = 0.2
        elif small_dt_flag == 0 and dV < 0.02:
            """when not close to end Voltage, increase the time step if the dV is judged to be too small"""
            dt = dt * 2.0
        elif small_dt_flag == 0 and dV > 0.02:
            """when not close to end Voltage but dV is too high, reduce time step"""
            dt = dt / 1.5
        elif small_dt_flag == 1 and dV > 0.005:
            """if close to end voltage and dV is too high, reduce time step"""
            dt = dt / 2.0
        elif small_dt_flag == 1 and dV < 0.001:
            """if close to end voltage and dV is too low, increase time step"""
            dt = dt * 1.5

        if dtpulse == 1:
            """pulse mode is on"""
            if pulseTime <= 10:
                dt = 0.3

            if pulseSchedule[1] == 1 and DOD >= pulseSchedule[2] / 1.03:
                dt = numpy.maximum(2.0, dt / 4.0)
            elif pulseSchedule[1] == 4 and capacity >= pulseSchedule[2] / 1.03:
                dt = numpy.maximum(2.0, dt / 5.0)
            elif pulseSchedule[1] == 1 and DOD >= pulseSchedule[2] / 1.25:
                dt = numpy.maximum(5.0, dt / 1.5)
            elif pulseSchedule[1] == 4 and capacity >= pulseSchedule[2] / 1.3:
                dt = numpy.maximum(5.0, dt / 4.0)
            elif pulseSchedule[1] == 2 and (pulseTime + dt) > pulseSchedule[2]:
                dt = pulseSchedule[2] - pulseTime
    else:
        """cccv mode"""
        targetIapp = IcutOff
        error = (targetIapp - Iapp) / targetIapp

        if error < 0 and numpy.abs(error) < 0.1 and realdt > 0.4:
            """closing in on end current"""
            # print("Reducing CV time step")
            dt = numpy.abs(error) * dt_fixed * 2.0
            if dt < 0.5:
                dt = 0.5
        elif error > 0:
            # print 'min time step', error
            dt = 1.0
        else:
            # print 'cccv dt fixed'
            dt = dt_fixed

    if step == 4 and stepTime + dt > trest:
        """make sure dt isn't too big for the OCV period"""
        dt = trest - stepTime
        if dt > trest / 5:
            dt = trest / 5

    return dt


def timeControlSlow(
    dt,
    dt_fixed,
    voltage,
    dV,
    Iapp,
    step,
    stepIterations,
    IcutOff,
    trest,
    stepTime,
    maxV,
    minV,
):
    """Control the time step. Either faster or slower based on situation."""

    realdt = dt

    if step != 3:
        """Find out if the voltage is closing in on the end voltage, if so, set a warning flag"""
        if Iapp < 0 and voltage < (minV + 0.01):
            small_dt_flag = 1
            # dt = 0.01
            dt = 0.05
        # dt = 0.02
        elif (Iapp < 0 and voltage < (minV + 0.1)) or (
            Iapp > 0 and voltage > (maxV - 0.002)
        ):
            small_dt_flag = 1
            # dt = 0.1
            dt = 0.5
        elif (Iapp < 0 and voltage < (minV + 0.25)) or (
            Iapp > 0 and voltage > (maxV - 0.1)
        ):
            small_dt_flag = 1
            dt = 10
        else:
            small_dt_flag = 0

        if stepIterations < 3:
            """for first couple of time steps in a cycle, keep dt small for good resolution in transition period"""
            dt = 0.1
        elif small_dt_flag == 0 and dV < 0.01:
            """when not close to end Voltage, increase the time step if the dV is judged to be too small"""
            dt = dt * 2.0
        elif small_dt_flag == 0 and dV > 0.01:
            dt = dt / 1.5
        elif small_dt_flag == 1 and dV > 0.005:
            dt = dt / 2.0
        elif small_dt_flag == 1 and dV < 0.001:
            dt = dt * 1.5
    else:
        """cccv mode"""
        # print 'cccv mode'
        targetIapp = IcutOff
        error = (targetIapp - Iapp) / targetIapp
        # print error,realdt
        if error < 0 and numpy.abs(error) < 0.1 and realdt > 0.9:
            # print("Reducing CV time step")
            dt = numpy.abs(error) * dt_fixed
            if dt < 1.0:
                dt = 1.0
        elif error > 0:
            # print 'min time step', error
            dt = 1.0
        else:
            # print 'cccv dt fixed'
            dt = dt_fixed

    if step == 4 and stepTime + dt > trest:
        """make sure dt isn't too big for the OCV period"""
        dt = trest - stepTime
        if dt > trest / 5:
            dt = trest / 5

    return dt


def arrhenius(phi_ref, Ea, T, T_ref=298.0):
    """return Arrhenius expression using reference values"""
    return phi_ref * numpy.exp(-Ea / 8.3145 * ((1.0 / T) - (1.0 / T_ref)))


class electrode:
    """Controlled by the class singleCell"""

    def __init__(self, parameters, electrodeType, cellNumber, writeData, T):
        """kinetic and transport parameters:"""

        def returnLength(parameter):
            try:
                x = len(parameter)
            except:
                x = 1
            return x

        # Solid phase diffusion coefficient
        self.Ds = {"A": 0, "Ea": 0}
        self.Ds["A"] = {2: lambda: parameters["Ds"][0], 1: lambda: parameters["Ds"]}[
            returnLength(parameters["Ds"])
        ]()
        self.Ds["Ea"] = {2: lambda: parameters["Ds"][1], 1: lambda: 0}[
            returnLength(parameters["Ds"])
        ]()

        # Charge transfer rate constant '
        self.kct = {"A": 0, "Ea": 0}
        self.kct["A"] = {2: lambda: parameters["kct"][0], 1: lambda: parameters["kct"]}[
            returnLength(parameters["kct"])
        ]()
        self.kct["Ea"] = {2: lambda: parameters["kct"][1], 1: lambda: 0}[
            returnLength(parameters["kct"])
        ]()

        # Side reaction exchange current density '
        self.i0s = {"A": 0, "Ea": 0}
        self.i0s["A"] = {2: lambda: parameters["i0s"][0], 1: lambda: parameters["i0s"]}[
            returnLength(parameters["i0s"])
        ]()
        self.i0s["Ea"] = {2: lambda: parameters["i0s"][1], 1: lambda: 0}[
            returnLength(parameters["i0s"])
        ]()

        self.soc = parameters["soc0"]
        self.soc0 = parameters["soc0"]
        # self.soc0_dch=parameters['soc0_dch']

        self.DsFactor = parameters["DsFactor"]

        # side reaction parameters
        self.Ms = parameters["Ms"]  # molar mass of sei
        self.rhos = parameters["rhos"]  # density of sei
        self.Lsei = {
            "present": parameters["Lsei"],
            "last_timeStep": parameters["Lsei"],
        }  # sei thickness
        self.Lsei0 = parameters["Lsei0"]  # sei thickness at t=0
        self.Erefs = parameters["Erefs"]  # Eref of side reaction
        self.ks = parameters["ks"]  # conductivity of sei
        self.Rsei = parameters["Rsei"]  # sei resistance

        # structural and design parameters:
        self.L = parameters["L"]
        self.capacity = parameters["capacity"]
        self.area = parameters["area"]
        self.apparentDensity = parameters["apparentDensity"]
        self.massFracAM = parameters["massFracAM"]
        self.massFracCarbon = parameters["massFracCarbon"]
        self.massFracBinder = 1.0 - self.massFracAM - self.massFracCarbon
        self.Rp = parameters["Rp"]
        self.electrodeType = electrodeType  # 1 for cathode, 2 for anode

        self.activeMaterialType = parameters["type"]
        self.cellNumber = cellNumber
        self.Eref_fudge = 0

        # calculated values
        self.volume = self.area * self.L
        self.porosity = self.porosity()
        self.volFracAM = self.volFracSolids(self.activeMaterialType)
        self.volFracCarbon = self.volFracSolids("carbon")
        self.volFracBinder = self.volFracSolids("PVDF")
        self.surfaceArea = self.surfaceArea()
        self.mass = self.volume * (
            density(self.activeMaterialType)
            * self.volFracSolids(self.activeMaterialType)
            * (1.0 - self.porosity)
            + density("carbon") * self.volFracSolids("carbon") * (1.0 - self.porosity)
            + density("PVDF") * self.volFracSolids("PVDF") * (1.0 - self.porosity)
            + density("electrolyte") * self.porosity
        )
        self.cmax = (
            self.capacity
            / 96485
            * 3.6
            * 100
            * 100
            * 100
            * density(self.activeMaterialType)
            / 1000
        )  # mol/m^3

        self.N = 25  # number of grid points for finite difference method
        self.socList = numpy.transpose([numpy.ones(self.N) * self.soc])
        self.socList_lastTimeStep = numpy.transpose([numpy.ones(self.N) * self.soc])

        self.J = {"J": 0.0, "Js": 0.0}
        self.J_last_timeStep = {"J": 0.0, "Js": 0.0}
        self.intJ = {"present": 0.0, "last_timeStep": 0.0}

        self.phi = {
            "present": Eref(self.activeMaterialType, self.soc, T, self.Eref_fudge),
            "old": Eref(self.activeMaterialType, self.soc, T, self.Eref_fudge),
            "last_timeStep": 0,
        }  # {'present' : 4.2, 'old' : 4.2}

        self.eta = 0
        self.Rp_val = 0

        # this is Ds(soc). the factor is a fitting parameter
        if self.activeMaterialType == "NMC":
            self.Ds["A"] = self.DsFactor * returnDs(self.soc, "NMC")

    def volFracSolids(self, species):
        """returns the volume fraction of species wrt solids"""
        totVolume = (
            self.massFracAM / density(self.activeMaterialType)
            + self.massFracCarbon / density("carbon")
            + self.massFracBinder / density("PVDF")
        )
        numerator = {
            self.activeMaterialType: lambda: self.massFracAM
            / density(self.activeMaterialType),
            "carbon": lambda: self.massFracCarbon / density("carbon"),
            "PVDF": lambda: self.massFracBinder / density("PVDF"),
        }[species]()
        return numerator / totVolume

    def porosity(self):
        """returns the porosity of the electrode"""
        return 1.0 - (self.apparentDensity * 1000.0) / (
            self.volFracSolids(self.activeMaterialType)
            * density(self.activeMaterialType)
            + self.volFracSolids("carbon") * density("carbon")
            + self.volFracSolids("PVDF") * density("PVDF")
        )

    def surfaceArea(self):
        """returns the surface area of the electrode in m^2"""
        return (
            (
                3.0
                * self.volFracSolids(self.activeMaterialType)
                * (1 - self.porosity)
                / self.Rp
            )
            * self.L
            * self.area
        )

    def locCurrent1(self, Iapp):
        """Returns the intercalaction current density in either anode or cathode"""
        changeSign = {2: lambda: -1.0, 1: lambda: 1.0}[self.electrodeType]()
        self.J_last_timeStep["J"] = self.J["J"]
        self.J["J"] = changeSign * Iapp / self.surfaceArea - self.J["Js"]
        return self.J["J"]  # changeSign*Iapp/self.surfaceArea

    def locCurrent(self, Iapp):
        changeSign = {2: lambda: -1.0, 1: lambda: 1.0}[self.electrodeType]()
        self.J["J"] = changeSign * Iapp / self.surfaceArea - self.J["Js"]
        return self.J["J"]

    def storeFDData(self, soc):
        self.socAll = soc

    def applySideReactionCorrection(self, theta, cycle):
        """reduces the concentration in the pos electrode by amount theta"""
        self.socAll = self.socAll - theta

    def finiteDifference(self, dt, Iapp, totTime, T):
        """
        Solves FD speherical diffusion equation:
        d(soc)/dt = D/Rp^2 * div(grad(soc)) + 2*D/x/Rp^2*grad(soc)
        Crank Nicholson (forward in time) scheme
        For cluster: remove all sparse matrices, use soc = numpy.mat(scipy.linalg.inv(A)*b)
        """

        J = self.locCurrent(Iapp)

        if totTime > 0:
            N = self.N  # number of grid points
            dx = 1 / (N - 1.0)  # grid spacing
            Ds = arrhenius(self.Ds["A"], self.Ds["Ea"], T)

            delta = -J * self.Rp / self.cmax / Ds / 96485.0
            K = Ds / self.Rp / self.Rp

            # grid points
            x = numpy.linspace(0, 1, N)

            # source term
            F = numpy.zeros([N, 1])
            F[-1] = (
                2.0 * K * dx * delta / (dx * dx) + 2.0 * K * delta / x[-1]
            )  # 2.0*delta*(1.0/dx+1.0/x[-1]/self.Ds)

            # create A matrix with boundary conditions
            # D1 is for second derivative: div(grad(C)) and D2 is for 1/x/D*grad(C)
            # D3 is the combined D1 and D2 matrices

            """ 
			I've read that spdiags is supposed to be faster than csr_matrix 
			but this doesn't seem to be the case 
			"""

            """
			data=numpy.ones((3,N))
			data[1]=-2*data[1]
			data[0,-2]=2;data[-1,1]=2
			diags=[-1,0,1]
			D1 = scipy.sparse.spdiags(data,diags,N,N)*K/(dx**2)

			data=numpy.zeros((3,N))
			data[2,2:]=1/x[1:-1]
			data[0,0:-2]=-1/x[1:-1]
			D2 = scipy.sparse.spdiags(data,diags,N,N)*2.0*K/(2.0*dx)
			
			#old way to generate D1/D2 matrix
			D1=numpy.zeros([N,N])
			D2=numpy.zeros([N,N])
			D1[0,0:2]=[-2,2] # zero gradient for D1
			D1[-1,-2:]=[2,-2]
			
			for i in range(1,N-1):
				D1[i,i-1:i-1+3] = [1.0,-2.0,1.0]
				D2[i,i-1:i-1+3] = [-1.0/x[i],0.0,1.0/x[i]] #x is the grid spacing
			"""

            """ new way to generate D1/D2 matrix"""
            D1 = numpy.eye(N) * -2 + numpy.eye(N, k=1) + numpy.eye(N, k=-1)
            D1[0, 1] = 2
            D1[-1, -2] = 2
            x1 = numpy.zeros([N])
            x1[2:] = 1 / x[1:-1]
            x2 = numpy.zeros([N])
            x2[: N - 2] = -1 / x[1:-1]
            D2 = numpy.eye(N, N, k=1) * x1 + numpy.eye(N, N, k=-1) * x2

            # gradient boundary condition for D1
            D1 = D1 * K / (dx**2)
            D2 = 2.0 * K * D2 / (2.0 * dx)
            D1 = scipy.sparse.csr_matrix(D1)
            D2 = scipy.sparse.csr_matrix(D2)

            try:
                I = scipy.sparse.identity(N)
            except AttributeError:
                I = numpy.eye(N)  # for the cluster

            A = I - dt / 2.0 * D1 - dt / 2.0 * D2
            b = (I + dt / 2.0 * D1 + dt / 2.0 * D2) * self.socList_lastTimeStep + dt * F

            try:
                socList = numpy.transpose(numpy.mat(scipy.sparse.linalg.spsolve(A, b)))
            except:
                socList = numpy.mat(scipy.linalg.inv(A) * b)  # cluster

            self.socList = numpy.array(socList)

        return self.socList

    def save_lastTimeStep(self):
        self.socList_lastTimeStep = self.socList
        # self.Iapp['last_timeStep'] = self.Iapp['present']
        self.J_last_timeStep["J"] = self.J["J"]
        self.J_last_timeStep["Js"] = self.J["Js"]
        self.intJ["last_timeStep"] = self.intJ["present"]
        self.phi["last_timeStep"] = self.phi["present"]
        self.Lsei["last_timeStep"] = self.Lsei["present"]

    def calc_intJ(self, dt, Iapp):
        J = self.locCurrent(Iapp)
        self.intJ["present"] = (
            self.intJ["last_timeStep"]
            + ((J / 96485.0 + self.J_last_timeStep["J"] / 96485.0) / 2.0) * dt
        )
        return self.intJ["present"]

    def polynomialApproximation(self, dt, Iapp, T):

        """
        Use the polynomial approximation to sovle the 1D spherical equation
        """

        J = self.locCurrent(Iapp)

        Ds = arrhenius(self.Ds["A"], self.Ds["Ea"], T)

        intJ = self.calc_intJ(dt, Iapp)
        c_avg = -3.0 / self.Rp * intJ + self.cmax * self.soc0  # c_avg_pos_last_timeStep
        c_surf = (-J / 96485.0 / 5.0 + c_avg * Ds / self.Rp) * self.Rp / Ds

        return c_surf / self.cmax

    def calcSOC(self, dt, Iapp, totTime, T, method="fd"):
        """
        Calculate the soc at time totTime
        Basically a wrapper for finiteDifference method
        """
        if method == "fd":
            soc = self.finiteDifference(dt, Iapp, totTime, T)
            self.soc = soc[-1][0]

        elif method == "pa":
            self.soc = self.polynomialApproximation(dt, Iapp, T)

        else:
            print("Error: Please select either fd or pa for soc solver")

    def butlerVolmer(self, Iapp, ce, alpha, T):
        """Uses BV expression to calc eta"""
        Rg = 8.3145
        F = 96485.0
        J = self.locCurrent(Iapp)
        soc = self.soc

        if soc > 1.0:
            soc = 0.99999
        elif soc < 0:
            soc = 0.00001

        # print("kct={0},soc={1},T={2}".format(arrhenius(self.kct['A'],self.kct['Ea'],T),soc,T))
        i0 = (
            F
            * arrhenius(self.kct["A"], self.kct["Ea"], T)
            * numpy.power(self.cmax - soc * self.cmax, 0.5)
            * numpy.power(soc * self.cmax, 0.5)
            * numpy.power(ce, 0.5)
        )

        C1 = J / (2.0 * i0)
        C2 = numpy.sqrt(J * J + 4.0 * i0 * i0) / (2.0 * i0)
        constant_1 = C1 - C2
        constant_2 = C1 + C2
        constant = numpy.maximum(constant_1, constant_2)
        if constant_1 / constant_2 > 0:
            print("Check butlerVolmer(): both constants are same sign")

        self.eta = Rg * T / F / alpha * numpy.log(constant)
        return self.eta

    def RpCalc(self, ce, T):
        """
        Calculates polarization resistance of electrode using di/deta ohm m^2 geometric area
        """

        Rg = 8.3145
        F = 96485.0
        alpha = 0.5
        soc = self.soc
        i0 = (
            F
            * arrhenius(self.kct["A"], self.kct["Ea"], T)
            * numpy.power(self.cmax - soc * self.cmax, 0.5)
            * numpy.power(soc * self.cmax, 0.5)
            * numpy.power(ce, 0.5)
        )
        # self.butlerVolmer(J, ce, soc, alpha, T)
        # self.butlerVolmer(J, ce, alpha, T)

        Y = i0 * (
            alpha * F / Rg / T * numpy.exp(alpha * F / Rg / T * self.eta)
            + alpha * F / Rg / T * numpy.exp(-alpha * F / Rg / T * self.eta)
        )
        self.Rp_val = 1.0 / Y + self.Rsei
        return self.Rp_val / self.surfaceArea * self.area

    def potential(self, Iapp, ce, alpha, T):
        """
        returns the electrode potential
        JR = J_neg (A/m^2 SA * ohm m^2 SA)
        """
        JRfilm = self.locCurrent(Iapp) * self.Rsei

        self.phi["old"] = self.phi["present"]
        self.phi["present"] = (
            self.butlerVolmer(Iapp, ce, alpha, T)
            + Eref(self.activeMaterialType, self.soc, T, self.Eref_fudge)
            + JRfilm
        )

    def ohmicResistance(self, Iapp, T, ce, brugg):
        """returns the ohmic resistance of the electrode"""
        return self.L / (
            2.0 * self.area * ionicConductivity(ce, T, brugg, self.porosity)
        )

    def sideReaction(self, Iapp, T, cycle):
        """
        Calculates the side reaction current density
        Units: A/m^2 (surface area basis)
        """

        F = 96485.0
        Rg = 8.3145
        alpha = 0.5

        Eref_side = self.Erefs
        i0 = arrhenius(self.i0s["A"], self.i0s["Ea"], T)
        JRfilm = (
            self.locCurrent(Iapp) * self.Rsei
        )  # -> should this be J*Rsei or Js*Rsei
        # note: in this case, i0 has units of A/m^2 surface area
        if Iapp > 0 and cycle > 1:
            Js = -i0 * numpy.exp(
                -alpha * F / Rg / T * (self.phi["present"] - Eref_side - JRfilm)
            )
        else:
            Js = 0

        self.J["Js"] = Js

    def Rfilm(self, dt):
        """
        Film resistance calculation
        Units: M=kg/mol, rho=kg/m^3, kp=S/m
        Units: Lfilm=m, Rfilm: ohm m^2 (surface area)
        """
        F = 96485.0

        # Lfilm = -Js*M/rho/F*dt+LfilmOld
        Lsei = -self.J["Js"] * self.Ms / self.rhos / F * dt + self.Lsei["last_timeStep"]
        Rsei = Lsei / self.ks

        self.Rsei = Rsei
        self.Lsei["present"] = Lsei

    def Iapp(self):
        """
        Returns the total current
        """
        return self.J["J"] * self.surfaceArea


class separatorOrFoil:
    """Controlled by the class singleCell"""

    def __init__(self, parameters, cellNumber, writeData):
        self.L = parameters["L"]
        self.porosity = parameters["porosity"]
        self.area = parameters["area"]
        self.volume = self.L * self.area
        self.mass = self.volume * (
            density(parameters["type"]) + density("electrolyte") * self.porosity
        )
        self.cellNumber = cellNumber

    def ohmicResistance(self, Iapp, T, ce, brugg):
        return self.L / (self.area * ionicConductivity(ce, T, brugg, self.porosity))


class singleCell:
    """
    The controller for the single particle model.
    This object represents an entire cell and controls the classes:
    electrode, separatorOrFoil, cycleSchedule, getInputs


    Keyword arguments:
    parameters_list -- string, the parameters file name, typically an xlsx
    maxCycles -- integer, the maximum number of cycles to run
    cellNumber -- integer, a unique cell number identifyer
    writeData -- integer, whether to write data or not


    """

    def __init__(self, parameters_list, maxCycles, cellNumber, writeData):

        xlInterface = True
        parametersFileName = ""
        cycleScheduleFileName = ""
        parameters_file = ""
        if len(parameters_list) == 1:
            parameters_file = parameters_list[0]
            xlInterface = True
        elif len(parameters_list) == 2:
            parameters_file = parameters_list[0]
            cycleScheduleFileName = parameters_list[1]
            xlInterface = False
        else:
            print("Wrong format for input parameters argument.")

        def returnLength(parameter):
            try:
                x = len(parameter)
            except:
                x = 1
            return x

        def returnSolver(string):
            return {True: lambda: "fd", False: lambda: "pa"}[
                string == "Finite difference"
            ]()

        self.schedule = cycleSchedule(cycleScheduleFileName, maxCycles, xlInterface)

        inputs = getInputs(xlInterface, parameters_file)
        pos_parms = inputs.positive
        neg_parms = inputs.negative
        sep_parms = inputs.separator
        Alfoil_parms = inputs.Alfoil
        Cufoil_parms = inputs.Cufoil
        other_parms = inputs.others
        self.pos_solver = inputs.pos_solver
        self.neg_solver = inputs.neg_solver

        if xlInterface:
            self.schedule.schedule = inputs.cycle.tolist()
        self.schedule.getIapp()

        self.T = other_parms["T"]
        self.Rint = 0.0  # this is set through self.Rint()

        print(self.schedule.Iapp)

        # Instantiate cell component objects:

        self.cathode = electrode(pos_parms, 1, cellNumber, writeData, self.T)
        self.anode = electrode(neg_parms, 2, cellNumber, writeData, self.T)
        self.sep = separatorOrFoil(sep_parms, cellNumber, writeData)
        self.Al_foil = separatorOrFoil(Alfoil_parms, cellNumber, writeData)
        self.Cu_foil = separatorOrFoil(Cufoil_parms, cellNumber, writeData)

        self.Tamb = self.T
        self.TcoolantOut = self.T
        self.Qheat = 0

        self.isothermal = {True: lambda: True, False: lambda: False}[
            other_parms["isothermal"] == "yes"
        ]()
        self.ce = other_parms["ce"]
        self.Cp = other_parms["Cp"]
        self.h = other_parms["h"]
        self.Aexposed = other_parms["Aexposed"]
        self.alpha = pos_parms["alpha"]
        self.maxVcell = other_parms["maxVcell"]
        self.minVcell = other_parms["minVcell"]
        self.IcutOff = other_parms["IcutOff"]

        self.electrolyteFactor = {"A": 0, "Ea": 0}
        self.electrolyteFactor["A"] = {
            2: lambda: other_parms["electrolyteFactor"][0],
            1: lambda: other_parms["electrolyteFactor"],
        }[returnLength(other_parms["electrolyteFactor"])]()
        self.electrolyteFactor["Ea"] = {
            2: lambda: other_parms["electrolyteFactor"][1],
            1: lambda: 0,
        }[returnLength(other_parms["electrolyteFactor"])]()

        self.totMass = (
            self.cathode.mass
            + self.sep.mass
            + self.anode.mass
            + self.Al_foil.mass
            + self.Cu_foil.mass
        )
        self.totVolume = (
            self.cathode.volume
            + self.sep.volume
            + self.anode.volume
            + self.Al_foil.volume
            + self.Cu_foil.volume
        )
        self.density = self.totMass / self.totVolume
        self.step = 1  # for individual cell charging, the step should be known [1 for discharge, 2 for charge, 3 for CCCV, 4 for OCV]
        self.cellNumber = cellNumber

        self.V = {"present": 4.2, "last_timestep": 4.2}
        self.capacity = {
            "cumulative_discharge": 0,
            "cumulative_charge": 0,
            "lastCycle_discharge": 0,
            "lastCycle_charge": 0,
            "nominal": 0,
        }
        self.energy = {
            "cumulative_discharge": 0,
            "cumulative_charge": 0,
            "lastCycle_discharge": 0,
            "lastCycle_charge": 0,
            "nominal": 0,
        }

        # self.calcCellVoltage() #initialize cell voltage at t=0

        self.IVList = numpy.zeros([1, 2])

    def calcCapacity(self):
        """calculate the cumulative capacity and energy in As"""

        Iapp = self.schedule.Iapp["present"]
        Iapp_old = self.schedule.Iapp["last_timestep"]
        V = self.V["present"]
        V_old = self.V["last_timestep"]

        if self.schedule.Iapp["present"] < 0:
            self.capacity["cumulative_discharge"] += (
                ((-1.0 * Iapp) + (-1.0 * Iapp_old)) / 2.0 * self.schedule.dt
            )
            self.energy["cumulative_discharge"] += (
                ((-1.0 * Iapp * V) + (-1.0 * Iapp_old * V_old)) / 2.0 * self.schedule.dt
            )
        elif self.schedule.Iapp["present"] > 0:
            self.capacity["cumulative_charge"] += (
                ((1.0 * Iapp) + (1.0 * Iapp_old)) / 2.0 * self.schedule.dt
            )
            self.energy["cumulative_charge"] += (
                ((1.0 * Iapp * V) + (1.0 * Iapp_old * V_old)) / 2.0 * self.schedule.dt
            )

    def resetCapacities(self):
        """reset capacities at the beginning of a new cycle"""

        if self.schedule.step == 0 and self.schedule.stepIterations == 0:
            if self.schedule.cycle == 2:
                self.capacity["nominal"] = self.capacity["cumulative_discharge"]
                self.energy["nominal"] = self.energy["cumulative_discharge"]

            self.capacity["lastCycle_discharge"] = self.capacity["cumulative_discharge"]
            self.capacity["lastCycle_charge"] = self.capacity["cumulative_charge"]
            self.energy["lastCycle_discharge"] = self.energy["cumulative_discharge"]
            self.energy["lastCycle_charge"] = self.energy["cumulative_charge"]

            self.capacity["cumulative_discharge"] = 0
            self.energy["cumulative_discharge"] = 0
            self.capacity["cumulative_charge"] = 0
            self.energy["cumulative_charge"] = 0

    def returnCumCapacity(self, data):
        """
        Returns the cumulative capacity where data is [time,current]
        Capacity in mAh, [discharge,charge]
        """

        tmp_dch = data * 1.0
        tmp_ch = data * 1.0

        for i in range(len(data[:, 1])):
            if tmp_dch[i, 1] > 0:
                tmp_dch[i, 1] = 0
            if tmp_ch[i, 1] < 0:
                tmp_ch[i, 1] = 0
        # note:initial keyword only available in scipy version 0.11+
        dch = cumtrapz(-tmp_dch[:, 1], tmp_dch[:, 0]) / 60 / 60 * 1000
        ch = cumtrapz(tmp_ch[:, 1], tmp_ch[:, 0]) / 60 / 60 * 1000

        # return numpy.transpose(numpy.array([dch,ch]))
        return [dch[-1], ch[-1]]

    def returnDepthOfCycle(self, cycle, refCapacityCycle=0):
        """
        Returns the depth of discharge (if discharging) or state of charge (if charging), relative to the previously stored charge capacity
        Notes: If refCapacityCycle=0, then the cumulative capacity will be divided by the charge capacity of the last cycle
        otherwise it will get divided by the charge capacity at the specified cycle number.
        """

        if refCapacityCycle == 0:
            referenceCapacity = cycle - 2
        else:
            referenceCapacity = refCapacityCycle - 1

        if self.step == 1:
            return self.timeData[-1, 16] / self.capacity[referenceCapacity, 2]
        else:
            return self.timeData[-1, 17] / self.capacity[referenceCapacity, 2]

    def calcQheat(self, Iapp):
        """
        Calculate the heat generation of the cell. This is primarily used to calc the increase
        in coolant temperature.
        """

        Cp = self.Cp
        V = self.totVolume
        rho = (self.cathode.mass + self.sep.mass + self.anode.mass) / (
            self.cathode.volume + self.sep.volume + self.anode.volume
        )
        dUdT_pos = dUdT(self.cathode.activeMaterialType, self.cathode.soc)
        dUdT_neg = dUdT(self.anode.activeMaterialType, self.anode.soc)
        Rohmic = self.Rohm(Iapp)
        Told = self.T
        eta_pos = self.cathode.eta
        eta_neg = self.anode.eta

        self.Qheat = Iapp * Told / V * (dUdT_pos - dUdT_neg) + Iapp / V * (
            eta_pos - eta_neg + Iapp * Rohmic
        )

    def calcTambient(self, TcoolantIn, flowRate):
        """
        Calculate the average ambient temp and the ambient temperature (ie. cooling fluid temp) at the edge,
        or outlet of the cell based on using the outlet temp of the previous cell as inlet temp.
        This is a simple Q=mCpdT calculation
        flowRate is in LPM
        Tinlet in K (but gets converted in deg C)
        """
        coolantDensity = (
            -0.0035 * TcoolantIn * TcoolantIn + 1.7938 * TcoolantIn + 768.62
        )  # desntiy of water (kg/m^3, T in K)
        coolantHeatCapacity = (
            3.7694718940962e04
            - 3.9808781886345e02 * TcoolantIn
            + 1.7743407597823 * TcoolantIn * TcoolantIn
            - 3.5210730266402e-03 * TcoolantIn * TcoolantIn * TcoolantIn
            + 2.6283157323435e-06 * TcoolantIn * TcoolantIn * TcoolantIn * TcoolantIn
        )  # Cp water (J/kg-K, T in K)

        m = flowRate / 1000.0 * coolantDensity / 60.0  # mass flow of coolant (kg/s)

        self.TcoolantOut = (
            self.Qheat * self.totVolume + m * coolantHeatCapacity * (TcoolantIn)
        ) / (
            m * coolantHeatCapacity
        )  # Tout in K
        # print TcoolantIn, self.Qheat*self.totVolume,m,coolantHeatCapacity,TcoolantIn,coolantDensity, self.TcoolantOut
        self.Tamb = (
            TcoolantIn + self.TcoolantOut
        ) / 2.0  # Tamb is the average of inlet/outlet Tamb

    def calcTemperature(self, Iapp, dt, totTime):
        """
        Solve a simple energy balance found in Eq 16 of Guo2011
        rho*V*Cp*dT/dt = IT*[dUpos/dT - dUneg/dT]+I*(eta_pos-eta_neg+I*R_ohm) - q

        """
        Cp = self.Cp
        V = self.totVolume
        rho = (self.cathode.mass + self.sep.mass + self.anode.mass) / (
            self.cathode.volume + self.sep.volume + self.anode.volume
        )
        dUdT_pos = dUdT(self.cathode.activeMaterialType, self.cathode.soc)
        dUdT_neg = dUdT(self.anode.activeMaterialType, self.anode.soc)
        Rohmic = self.Rohm(Iapp)
        h = self.h
        Aexposed = self.Aexposed
        Told = self.T
        Tamb = self.Tamb
        eta_pos = self.cathode.eta
        eta_neg = self.anode.eta

        numerator = (
            Cp * rho * Told * V
            + dt * Rohmic * Iapp * Iapp
            + (dt * eta_pos - dt * eta_neg) * Iapp
            + dt * h * Tamb * Aexposed
        )
        denominator = (
            Cp * rho * V + (dt * dUdT_neg - dt * dUdT_pos) * Iapp + dt * h * Aexposed
        )
        T = numerator / denominator

        # print("Cellnumber={0}, T={1}, Told={2}, dt={3}, eta_pos={4}, eta_neg={5}, Iapp={6}, Tamb={7}, soc_pos={8}, soc_neg={9}, totTime={10}".format(cellNumber,T,Told,dt,eta_pos,eta_neg,Iapp,Tamb,soc_pos,soc_neg,totTime))
        # print("cellNumber={0},T={1}".format(cellNumber,T))

        """
		#4th order runge-kutta
		T1 = Told
		K1=(Iapp*(eta_pos-eta_neg)+Iapp*T1*(dUdT_pos-dUdT_neg)+Iapp*Iapp*Rohmic+h*Aexposed*(Tamb-T1))/rho/Cp/V
		T2 = Told+K1*dt/2.0
		K2= (Iapp*(eta_pos-eta_neg)+Iapp*T2*(dUdT_pos-dUdT_neg)+Iapp*Iapp*Rohmic+h*Aexposed*(Tamb-T2))/rho/Cp/V
		T3 = Told+K2*dt/2.0
		K3=(Iapp*(eta_pos-eta_neg)+Iapp*T3*(dUdT_pos-dUdT_neg)+Iapp*Iapp*Rohmic+h*Aexposed*(Tamb-T3))/rho/Cp/V
		T4 = Told+K3*dt
		K4=(Iapp*(eta_pos-eta_neg)+Iapp*T4*(dUdT_pos-dUdT_neg)+Iapp*Iapp*Rohmic+h*Aexposed*(Tamb-T4))/rho/Cp/V

		T = Told + (1/6)*(K1+2*K2+2*K3+K4)*dt
		"""
        if T > 60 + 273.15:
            print("Warning: T=", T - 273.15, "degC")
        self.T = T

    def Rohm(self, Iapp):
        """
        Returns the ohmic resistance of the cell.
        The electrolyte factor is a fitting parameter meant especially for low temp fitting
        """
        return (
            self.cathode.ohmicResistance(Iapp, self.T, self.ce, 1.5)
            / arrhenius(
                self.electrolyteFactor["A"], self.electrolyteFactor["Ea"], self.T
            )
            + self.anode.ohmicResistance(Iapp, self.T, self.ce, 1.5)
            / arrhenius(
                self.electrolyteFactor["A"], self.electrolyteFactor["Ea"], self.T
            )
            + self.sep.ohmicResistance(Iapp, self.T, self.ce, 1.5)
            / arrhenius(
                self.electrolyteFactor["A"], self.electrolyteFactor["Ea"], self.T
            )
        )

    def calcCellVoltage(self):
        """The main algorithm. Calculate the cell voltage for the next time step"""

        def V_cell(Iapp, cycle, totTime, dt):

            # calc SEI film thickness and resistance
            self.anode.Rfilm(dt)
            self.cathode.Rfilm(dt)

            # calc side reaction current density
            self.anode.sideReaction(Iapp, self.T, cycle)
            self.cathode.sideReaction(Iapp, self.T, cycle)

            # this is Ds(soc). the factor is a fitting parameter
            if self.cathode.activeMaterialType == "NMC":
                self.cathode.Ds["A"] = self.cathode.DsFactor * returnDs(
                    self.cathode.soc, "NMC"
                )

            self.cathode.calcSOC(
                dt, Iapp, totTime, self.T, method=self.pos_solver
            )  # calc soc cathode
            self.anode.calcSOC(
                dt, Iapp, totTime, self.T, method=self.neg_solver
            )  # calc soc anode

            # ---------------

            VoldIt = 0
            V = 100
            its = 0
            doneInternalIts = 0

            while doneInternalIts == 0:
                """make sure that the internal iterations have converged"""
                its += 1

                self.cathode.potential(
                    Iapp, self.ce, self.alpha, self.T
                )  # calc phi cathode
                self.anode.potential(
                    Iapp, self.ce, self.alpha, self.T
                )  # calc phi anode

                VoldIt = V
                V = (
                    self.cathode.phi["present"]
                    - self.anode.phi["present"]
                    + Iapp * self.Rohm(Iapp)
                )
                # if (cycle == 4): print("t={0}, V={1}").format(totTime,V)
                dV_it = numpy.abs(V - VoldIt)
                if dV_it < 1e-4:
                    doneInternalIts = 1
                elif its > 10:
                    print(
                        f"Warning: Internal voltage iterations are at {its} and dV is {dV_it} but moving on."
                    )
                    doneInternalIts = 1

            if numpy.isnan(V):
                print("Error: You're getting NaNs!")
                sys.exit(0)

            return V

        def cv_residual(Iapp, Vtarget, cycle, totTime, dt):
            """
            Return the residual for constant voltage current search
            """
            V = V_cell(Iapp, cycle, totTime, dt)
            sse = (Vtarget - V) ** 2.0
            return sse

        dt = self.schedule.dt
        totTime = self.schedule.totTime
        cycle = self.schedule.cycle

        cvIteration = 1  # flag for constant voltage mode

        self.V["last_timeStep"] = self.V[
            "present"
        ]  # set last V here, otherwise cv iterations will mess it up
        self.resetCapacities()

        while cvIteration == 1:

            if self.schedule.mode == "cc":
                Iapp = self.schedule.Iapp["present"]
            elif self.schedule.mode == "cv":
                self.cvGuess()
                Iapp = self.schedule.Iapp["present"]

            # ---------------
            # Technically this should all be moved into the while loop.
            # Check to see the impact. Keeping out of the loop is faster since
            # soc only has to be calculated once.

            if not self.isothermal:
                self.calcTemperature(Iapp, dt, totTime)

            # Calculate the internal cell resistance
            self.Rinternal(Iapp)

            self.V["present"] = V_cell(Iapp, cycle, totTime, dt)

            if (
                self.schedule.mode == "cv"
                and numpy.abs(
                    self.V["present"] - self.schedule.schedule[self.schedule.step][1]
                )
                < 1e-3
            ):
                cvIteration = 0
            elif self.schedule.mode == "cc":
                cvIteration = 0

        # done cv iterations -> move to next time step
        self.calcQheat(Iapp)
        self.cathode.save_lastTimeStep()
        self.anode.save_lastTimeStep()
        self.calcCapacity()

    def cvGuess(self):
        """Guess current for constant voltage mode"""

        def numpyPop(x1, x2, numpyArray):
            """Append [x1,x2] to numpyArray and delete row 0"""
            numpyArray = numpy.vstack((numpyArray, [x1, x2]))
            if numpy.sum(numpyArray[0, :] == 0):
                numpyArray = numpy.delete(numpyArray, 0, axis=0)
            return numpyArray

        Iapp = self.schedule.Iapp["present"]
        V = self.V["present"]

        self.IVList = numpyPop(Iapp, V, self.IVList)

        Vtarget = self.schedule.schedule[self.schedule.step][1]

        if numpy.shape(self.IVList)[0] < 2:
            # List not long enough to interpolate
            if self.IVList[-1, 1] > Vtarget:
                Iapp = Iapp / 1.001
            else:
                Iapp = Iapp * 1.001
        else:
            # List should be long enough to interpolate
            # Ig = I1+(Vtar-V1)*(I2-I1)/(V2-V1)
            V1 = self.IVList[-2, 1]
            V2 = self.IVList[-1, 1]
            I1 = self.IVList[-2, 0]
            I2 = self.IVList[-1, 0]

            Iapp = I1 + (Vtarget - V1) * ((I2 - I1) / (V2 - V1))

        # print("New guess for CV mode: I={0}".format(Iapp))
        self.schedule.Iapp["present"] = Iapp

    def Rinternal(self, Iapp):
        """Return the internal resistance of the cell in Ohm"""

        self.Rint = (
            self.anode.RpCalc(self.ce, self.T)
            + self.cathode.RpCalc(self.ce, self.T)
            + self.Rohm(Iapp)
        )


class cycleSchedule:
    def __init__(self, scheduleFileName, maxCycles, xlInterface):

        self.schedule = []
        if not xlInterface:
            self.parseCycleScheduleFile(scheduleFileName)
        self.step = 0
        self.cycle = 0
        self.last_cycle = 0
        self.Iapp = {"present": 0, "last_timestep": 0}
        if not xlInterface:
            self.getIapp()  # this needs to be held off until self.schedule is populated
        self.maxCycles = maxCycles
        self.dt = 0.1
        self.stepIterations = 0
        self.stepTime = 0
        self.totTime = 0
        self.mode = "cc"  # cc or cv

    def parseCycleScheduleFile(self, fileName):
        """
        Reads cycle schedule from a file in the format:
        step[#] = [step_type, step_condition, stop_type, stop_condition, maxdt]
        step_type: 0 -> cc, 1 -> cv
        step_condition: current || voltage
        stop_type: 0->voltage condition, 1->DOD condition, 2->time condition, 3->current condition, 4->capacity condition
        stop_condition: voltage, DOD, time, current, capacity
        maxdt: maximum dt for this step
        """

        f = open(fileName, "r")
        schedule = []
        for line in f:
            line = line.strip()
            if re.search(r"^#", line) or line == "" or len(line.split(" ")) < 5:
                pass
            else:
                schedule.append(conv_numbers(line.split(" ")))

        f.close()

        self.schedule = schedule

    def getIapp(self):
        """Return the current"""
        if self.schedule[self.step][0] == 0:
            # self.Iapp['last_timestep'] = self.Iapp['present']
            self.Iapp["present"] = self.schedule[self.step][1]
        else:
            # self.Iapp['last_timestep'] = self.Iapp['present']
            self.Iapp["present"] = 0

    def checkStopCondition(self, voltage):
        """
        Algorithm is run after every time step

        Checks to see if we should switch the current

        step[#] = [step_type, step_condition, stop_type, stop_condition, maxdt]
        step_type: 0 -> cc, 1 -> cv
        step_condition: current || voltage
        stop_type: 0->voltage condition, 1->DOD condition, 2->time condition, 3->current condition, 4->capacity condition
        stop_condition: voltage, DOD, time, current, capacity
        maxdt: maximum dt for this step
        """
        self.last_cycle = self.cycle
        self.stepIterations += 1
        self.Iapp["last_timestep"] = self.Iapp["present"]

        if self.schedule[self.step][2] == 0 and (
            (self.Iapp["present"] < 0 and voltage <= self.schedule[self.step][3])
            or (self.Iapp["present"] > 0 and voltage >= self.schedule[self.step][3])
        ):
            "voltage condition met"
            self.advanceStep()

        elif (
            self.schedule[self.step][2] == 2.0
            and self.stepTime >= self.schedule[self.step][3]
        ):
            "time condition met"
            self.advanceStep()

        elif (
            self.schedule[self.step][2] == 3.0
            and self.Iapp["present"] <= self.schedule[self.step][3]
        ):
            "current condition met"
            self.advanceStep()

    def advanceStep(self):
        """
        Advance to the next step in the cycle
        Normally this happens after checking the stop conditions in the cycle, but
        it can also be forced. This may be needed when battery packs are simulated
        to maintain voltage levels of individual cells.
        """

        maxSteps = numpy.shape(self.schedule)[0]

        self.step = {False: lambda: self.step + 1, True: lambda: 0}[
            self.step == maxSteps - 1
        ]()
        # self.cycle = {False: lambda: self.cycle, True: lambda: self.cycle + 1}[self.step == maxSteps - 1]()
        self.cycle = {False: lambda: self.cycle, True: lambda: self.cycle + 1}[
            self.step == 0
        ]()

        self.stepIterations = 0
        self.stepTime = 0

        # set the next current
        if self.schedule[self.step][0] == 0:
            # next step is a constant current step
            self.Iapp["present"] = self.schedule[self.step][1]
            self.mode = "cc"
        elif self.schedule[self.step][0] == 1:
            # next step is a constant voltage step
            self.mode = "cv"

        print(f"cycle: {self.cycle}\tstep: {self.step}\t{self.schedule[self.step]}")
        # print("old cycle {0}, new cycle {1}").format(self.last_cycle, self.cycle)

    def set_dt(self, V, force=0):
        """
        Set the time step for the next iteration, but keep it within maxdt
        If force is specified, then force dt to equal that value

        step[#] = [step_type, step_condition, stop_type, stop_condition, maxdt]
        step_type: 0 -> cc, 1 -> cv
        step_condition: current || voltage
        stop_type: 0->voltage condition, 1->DOD condition, 2->time condition, 3->current condition, 4->capacity condition
        stop_condition: voltage, DOD, time, current, capacity
        maxdt: maximum dt for this step
        """

        if force == 0:
            maxdt = self.schedule[self.step][4]
            dt_old = self.dt

            self.dt = {False: lambda: self.dt * 1.05, True: lambda: maxdt}[
                self.dt >= maxdt
            ]()

            if (
                self.schedule[self.step][2] == 0
                and self.Iapp["present"] < 0
                and numpy.abs(V - self.schedule[self.step][3]) < 0.35
            ):
                self.dt = {False: lambda: dt_old / 1.5, True: lambda: 0.1}[
                    dt_old / 1.10 < 0.1
                ]()
            elif (
                self.schedule[self.step][2] == 0
                and self.Iapp["present"] > 0
                and numpy.abs(V - self.schedule[self.step][3]) < 0.25
            ):
                self.dt = {False: lambda: dt_old / 1.10, True: lambda: 0.1}[
                    dt_old / 1.10 < 0.1
                ]()
            elif self.schedule[self.step][2] == 2.0 and (
                self.stepTime + self.dt > self.schedule[self.step][3]
            ):
                self.dt = self.schedule[self.step][3] - self.stepTime

            if self.stepIterations < 5:
                self.dt = 0.1
        else:
            self.dt = force

    def advanceTime(self):
        self.stepTime += self.dt
        self.totTime += self.dt


class getInputs:
    """
    Grab input data from an excel spreadsheet using the openpyxl module

    Usage:
    inputs = getInputs("input_data.xlsx")
    parameters: inputs.positive,inputs.negative,inputs.Alfoil,inputs.Cufoil,inputs.Others,inputs.Separator
    """

    def __init__(self, xlInterface, input_file):

        if xlInterface:
            print("using excel interface")
            self.workbook = input_file
            self.worksheet = ["parameters", "cycle"]

            search_string = [
                "Positive",
                "Negative",
                "Al Foil",
                "Cu Foil",
                "Others",
                "Separator",
                "Cycle Conditions",
            ]

            self.positive = self.getExcelParameters(search_string[0], self.worksheet[0])
            self.negative = self.getExcelParameters(search_string[1], self.worksheet[0])
            self.Alfoil = self.getExcelParameters(search_string[2], self.worksheet[0])
            self.Cufoil = self.getExcelParameters(search_string[3], self.worksheet[0])
            self.others = self.getExcelParameters(search_string[4], self.worksheet[0])
            self.separator = self.getExcelParameters(
                search_string[5], self.worksheet[0]
            )

            # soc solver requires either pa or fd
            self.pos_solver = {True: lambda: "pa", False: lambda: "fd"}[
                self.positive["method"] == "Polynomial approximation"
            ]()
            self.neg_solver = {True: lambda: "pa", False: lambda: "fd"}[
                self.negative["method"] == "Polynomial approximation"
            ]()
            self.cycle = self.getCycleSchedule(search_string[6], self.worksheet[1])

        else:
            """If using a csv file instead of excel interface"""
            print(
                "using csv interface\nWarning: polynomial approximation is hard coded here."
            )
            csv_parameters = self.getCSVParameters(input_file)
            self.positive = csv_parameters["@pos"]
            self.negative = csv_parameters["@neg"]
            self.separator = csv_parameters["@sep"]
            self.Alfoil = csv_parameters["@Alfoil"]
            self.Cufoil = csv_parameters["@Cufoil"]
            self.others = csv_parameters["@others"]
            self.pos_solver = "pa"
            self.neg_solver = "pa"

    def findCells(self, cells, search_string):
        """
        Returns the row and col numbers of the search_string in an xls document
        cells is an openpyxl object

        Example: findCells(cells,"hot input") = [4,1]
        """

        maxRows = numpy.shape(cells)[0]
        maxCols = numpy.shape(cells)[1]

        col = 0
        row = 0
        for i in range(maxRows):
            for j in range(maxCols):
                if cells[i][j].value == search_string:
                    col = j
                    row = i
                    return [row, col]

    def convertToSIunits(self, dic_values, dic_units):
        """
        Simple routine to convert common units into SI units
        """
        for keys in dic_units:
            try:
                dic_values[keys] = {
                    "mm": lambda: dic_values[keys] / 1000.0,
                    "in": lambda: dic_values[keys] * 25.4 / 1000.0,
                    "um": lambda: dic_values[keys] / 1.0e6,
                    "mm^2": lambda: dic_values[keys] / 1000.0 / 1000.0,
                    "in^2": lambda: dic_values[keys] * 25.4 / 1000.0 * 25.4 / 1000.0,
                    "um^2": lambda: dic_values[keys] / 1.0e6 / 1.0e6,
                    "C": lambda: dic_values[keys] + 273.15,
                }[dic_units[keys]]()
            except:
                pass

        return dic_values

    def returnInputs(self, cells, start_index):
        """
        Returns a dictionary of hot_side and cold_side inputs from the input_data.xlsx file
        """

        def returnFloat(val):
            try:
                return float(val)
            except:
                return val

        startRow = start_index[0] + 2
        startCol = start_index[1] + 2
        values = []
        keys = []
        units = []

        max_num_keys = 50

        for i in range(max_num_keys):
            try:
                keys.append(cells[startRow + i][startCol].value.replace(" ", "_"))
                if cells[startRow + i][startCol + 2].value == None:
                    val = returnFloat(cells[startRow + i][startCol + 1].value)
                else:
                    val = returnFloat(
                        [
                            cells[startRow + i][startCol + 1].value,
                            cells[startRow + i][startCol + 2].value,
                        ]
                    )

                values.append(val)
                units.append(cells[startRow + i][startCol + 3].value)
            except IndexError:
                break
            except AttributeError:
                break

        parms = dict(zip(keys, values))
        # parms["fit_Nu"]=None
        units = dict(zip(keys, units))
        # units["fit_Nu"]=""

        return self.convertToSIunits(parms, units)

    def getExperimentalData(self, side, worksheet):
        """
        Get experimental data from excel file
        """
        import openpyxl as xl

        max_num_data = 100

        data = []

        wb = xl.load_workbook(filename=self.workbook)
        sheet = wb.get_sheet_by_name(name=worksheet)
        cells = [item for item in sheet.rows]

        id = self.findCells(cells, side)
        startRow = id[0] + 3
        startCol = id[1]

        for i in range(max_num_data):
            try:
                if cells[startRow + i][startCol].value == None:
                    break
                d1 = []
                [d1.append(cells[startRow + i][startCol + j].value) for j in range(7)]
                data.append(d1)
            except IndexError:
                break
            except AttributeError:
                break

        return numpy.array(data)

    def getExcelParameters(self, search_string, worksheet):

        wb = xl.load_workbook(filename=self.workbook, data_only=True)
        sheet = wb[worksheet]
        cells = [item for item in sheet.rows]

        return self.returnInputs(cells, self.findCells(cells, search_string))

    def getCycleSchedule(self, side, worksheet):
        """
        Get cycle schedule from excel file
        """

        data = []

        findCols = [
            "Step Type Index",
            "Step Condition",
            "Stop Type Index",
            "Stop Condition",
            "Max Time Step",
        ]

        cols = []

        wb = xl.load_workbook(filename=self.workbook, data_only=True)
        sheet = wb[worksheet]
        cells = [item for item in sheet.rows]

        max_num_data = numpy.shape(cells)[0]

        for i in range(numpy.shape(cells)[1]):
            if cells[1][i].value in findCols:
                cols.append(i - 1)

        id = self.findCells(cells, side)
        startRow = id[0] + 2
        startCol = id[1] + 1

        for i in range(max_num_data):
            try:
                if cells[startRow + i][startCol].value == None:
                    break
                d1 = []
                [d1.append(cells[startRow + i][startCol + j].value) for j in cols]
                data.append(d1)
            except IndexError:
                break
            except AttributeError:
                break

        return numpy.array(data)

    def getCSVParameters(self, fileName):
        """
        Reads simulation parameters from a file where each group of parameters begins with @label
        and the # is interpreted as a comment
        """
        # fileName = 'config.dat'
        f = open(fileName, "r")
        keys1 = []
        keys2 = []
        values = []
        dict1 = {}
        dict2 = {}
        for line in f:
            line = line.strip()
            if re.search(r"^#", line) or line == "":
                pass
            elif re.search(r"^@", line):
                dict2 = {}
                keys1.append(line.split("#")[0])
            else:
                keys2.append(line.split("=")[0])
                values.append(
                    line.split("#")[0].split("=")[-1].lstrip().rstrip().split(" ")
                )
                dict2[keys2[-1]] = conv_numbers(values[-1])
                dict1[keys1[-1]] = dict2

        f.close()

        return dict1


# ---------------------------------------------------------------------
def main():
    """
    Run the single particle model algorithms

    Creates the object singleCell and runs through the time stepping while saving data

    Inputs that hard coded
    parameters_list -> the file that contains the parameters, usually an xlsx file
    maxCycles -> the number of cycles to run
    writeData -> save data
    """

    maxCycles = 4

    colNames = "cycle,step,totTime_s,stepTime_s,current_A,voltage_V,dCap_As,cCap_As,posSOC,negSOC,Temp_K,Qheat_Wm3,Rint_ohm"

    parameters_list = ["supporting_files/parameters.xlsx"]

    # create the data directory
    data_dir = pathlib.Path().cwd() / "data"
    data_dir.mkdir(parents=True, exist_ok=True)

    cell1 = singleCell(parameters_list, maxCycles=maxCycles, cellNumber=1, writeData=1)
    # schedule = cycleSchedule("initial_steps.dat", 2)

    V = cell1.V
    print(f"Starting condition: {V}")
    time = 0
    run = 1
    data = 1

    while run == 1:
        cell1.schedule.advanceTime()

        # write data to file after each cycle
        if cell1.schedule.cycle > cell1.schedule.last_cycle:
            print(f"saving data for cycle {cell1.schedule.last_cycle}")
            file_name = data_dir / f"cell1_{cell1.schedule.last_cycle}.csv"
            saveData(file_name, data, headers=colNames)
            data = None

        cell1.calcCellVoltage()

        # save this data to file for every time step
        add_data = [
            cell1.schedule.cycle,
            cell1.schedule.step,
            cell1.schedule.totTime,
            cell1.schedule.stepTime,
            cell1.schedule.Iapp["present"],
            cell1.V["present"],
            cell1.capacity["cumulative_discharge"],
            cell1.capacity["cumulative_charge"],
            cell1.cathode.soc,
            cell1.anode.soc,
            cell1.T,
            cell1.Qheat,
            cell1.Rint,
        ]

        data = storeData(data, add_data)
        cell1.schedule.checkStopCondition(cell1.V["present"])
        cell1.schedule.set_dt(cell1.V["present"])
        # check for end of simulation
        run = {True: lambda: 0, False: lambda: 1}[
            cell1.schedule.cycle > cell1.schedule.maxCycles
        ]()

    print("Saving last cycle to data directory")
    file_name = data_dir / f"cell1_{cell1.schedule.last_cycle}.csv"
    saveData(file_name, data, headers=colNames)


if __name__ == "__main__":
    start_time = time.time()
    main()
    print(time.time() - start_time, " seconds")
