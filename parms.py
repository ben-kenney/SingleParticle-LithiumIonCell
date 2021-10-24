#!/usr/bin/python
import openpyxl as xl
import numpy

class getInputs:
	'''
	Grab input data from an excel spreadsheet using the openpyxl module
	
	Usage:
	inputs = getInputs("input_data.xlsx")
	parameters: inputs.positive,inputs.negative,inputs.Alfoil,inputs.Cufoil,inputs.Others,inputs.Separator
	'''
	
	def __init__(self,workbook):
		self.workbook = workbook
		self.worksheet = ["parameters", "cycle"]
		
		search_string = ["Positive","Negative","Al Foil","Cu Foil","Others","Separator","Cycle Conditions"]
		
		self.positive = self.getParameters(search_string[0],self.worksheet[0])
		self.negative = self.getParameters(search_string[1],self.worksheet[0])
		self.Alfoil = self.getParameters(search_string[2],self.worksheet[0])
		self.Cufoil = self.getParameters(search_string[3],self.worksheet[0])
		self.others = self.getParameters(search_string[4],self.worksheet[0])
		self.separator = self.getParameters(search_string[5],self.worksheet[0])
		self.cycle = self.getCycleSchedule(search_string[6],self.worksheet[1])

	def findCells(self,cells,search_string):
		'''
		Returns the row and col numbers of the search_string in an xls document
		cells is an openpyxl object
	
		Example: findCells(cells,"hot input") = [4,1]
		'''
	
		maxRows = numpy.shape(cells)[0]
		maxCols = numpy.shape(cells)[1]

	
		col=0;row=0
		for i in range(maxRows):
			for j in range(maxCols):
				if (cells[i][j].value == search_string):
					col=j;row=i
					return [row,col] 

	def convertToSIunits(self,dic_values,dic_units):
		'''
		Simple routine to convert common units into SI units	
		'''
		for keys in dic_units:
			try:
				dic_values[keys] = {
				'mm': lambda : dic_values[keys]/1000.0,
				'in': lambda : dic_values[keys]*25.4/1000.0,
				'um': lambda : dic_values[keys]/1.0e6,
				'mm^2': lambda : dic_values[keys]/1000.0/1000.0,
				'in^2': lambda : dic_values[keys]*25.4/1000.0*25.4/1000.0,
				'um^2' : lambda : dic_values[keys]/1.0e6/1.0e6,
				'C': lambda : dic_values[keys]+273.15}[dic_units[keys]]()
			except:
				pass

		return dic_values	
	
	def returnInputs(self,cells,start_index):
		'''
		Returns a dictionary of hot_side and cold_side inputs from the input_data.xlsx file
		'''
		
		def returnFloat(val):
			try:
				return float(val)
			except:
				return val
		
		startRow = start_index[0]+2
		startCol = start_index[1]+2
		values = []
		keys = []
		units = []
	
		max_num_keys = 50
	
		for i in range(max_num_keys):
			try:
				keys.append(cells[startRow+i][startCol].value.replace(" ","_"))
				if (cells[startRow+i][startCol+2].value == None):
					val = returnFloat(cells[startRow+i][startCol+1].value)
				else:
					val = returnFloat([cells[startRow+i][startCol+1].value,cells[startRow+i][startCol+2].value])
				
				values.append(val)
				units.append(cells[startRow+i][startCol+3].value)
			except IndexError:
				break
			except AttributeError:
				break

		parms = dict(zip(keys,values))
		#parms["fit_Nu"]=None
		units = dict(zip(keys,units))
		#units["fit_Nu"]=""
	
		return self.convertToSIunits(parms,units)
		
	def getExperimentalData(self,side,worksheet):
		'''
		Get experimental data from excel file
		'''	
		import openpyxl as xl

		max_num_data = 100
		
		data = []
		
		wb = xl.load_workbook(filename = self.workbook)
		sheet = wb.get_sheet_by_name(name = worksheet)
		cells = sheet.rows		
		
		id = self.findCells(cells,side)
		startRow = id[0]+3
		startCol = id[1]
		
		for i in range(max_num_data):
			try:
				if (cells[startRow+i][startCol].value == None): break				
				d1=[]
				[d1.append(cells[startRow+i][startCol+j].value) for j in range(7)]
				data.append(d1)		
			except IndexError:
				break
			except AttributeError:
				break
		
		return numpy.array(data)
		
	def getParameters(self,search_string,worksheet):

		wb = xl.load_workbook(filename = self.workbook)
		sheet = wb.get_sheet_by_name(name = worksheet)
		cells = sheet.rows
	
		return self.returnInputs(cells,self.findCells(cells,search_string))

	def getCycleSchedule(self,side,worksheet):
		'''
		Get cycle schedule from excel file
		'''	

		data = []
		cols = [0,1,3,4,5]
		
		wb = xl.load_workbook(filename = self.workbook)
		sheet = wb.get_sheet_by_name(name = worksheet)
		cells = sheet.rows		
		
		max_num_data = numpy.shape(cells)[0]
		
		
		id = self.findCells(cells,side)
		startRow = id[0]+2
		startCol = id[1]+1
		
		for i in range(max_num_data):
			try:
				if (cells[startRow+i][startCol].value == None): break				
				d1=[]
				[d1.append(cells[startRow+i][startCol+j].value) for j in cols]
				data.append(d1)		
			except IndexError:
				break
			except AttributeError:
				break
		
		return numpy.array(data)




def main():
	inputs = getInputs("parameters.xlsx")
	print inputs.positive
	print inputs.negative
	
if __name__ == "__main__":
	main()